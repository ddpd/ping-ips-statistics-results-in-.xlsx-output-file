import os
import platform
import subprocess
import time
from statistics import mean
from concurrent.futures import ThreadPoolExecutor, as_completed
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

SERVERS_DIR = 'servers'
RESULTS_DIR = 'ping_results'
PING_COUNT = 40
TIMEOUT_MS = 400
MAX_THREADS = 8
DEBUG = True

def setup_dirs():
    os.makedirs(SERVERS_DIR, exist_ok=True)
    os.makedirs(RESULTS_DIR, exist_ok=True)
    if DEBUG: print(f"Directories setup complete")

def debug_print(message):
    if DEBUG: print(f"DEBUG: {message}")

def execute_ping(host):
    param = '-n' if platform.system().lower() == 'windows' else '-c'
    command = ['ping', param, str(PING_COUNT)]

    ###########################################################################################
    if platform.system().lower() == 'windows':
        command.extend(['-w', str(TIMEOUT_MS)])  # Windows uses ms by default
    else:
        command.extend(['-W', str(TIMEOUT_MS / 1000)])  # Convert ms to seconds for Unix
    ###########################################################################################
    
    command.append(host)
    return subprocess.run(command, stdout=subprocess.PIPE, stderr=subprocess.PIPE, text=True)

def analyze_ping(output, host):
    output = output.lower()
    stats = {
        'times': [],
        'received': 0,
        'lost': PING_COUNT,
        'error': None,
        'raw_output': output
    }

    if 'received =' in output:
        stats['received'] = int(output.split('received = ')[1].split(',')[0])
    elif 'packets transmitted' in output:
        parts = output.split('packets transmitted, ')[1].split(' received,')
        stats['received'] = int(parts[0])
    
    stats['lost'] = PING_COUNT - stats['received']

    for line in output.split('\n'):
        if 'time=' in line:
            try:
                time_part = line.split('time=')[1].split()[0].replace('ms','')
                stats['times'].append(float(time_part))
            except (ValueError, IndexError) as e:
                debug_print(f"Time parse error: {e} in line: {line.strip()}")
    
    return stats

def ping_host(host):
    start_time = time.time()
    try:
        result = execute_ping(host)
        if result.returncode != 0:
            debug_print(f"Non-zero exit code ({result.returncode}) for {host}")
            return error_result(host, f"Ping failed (code {result.returncode})")
        
        stats = analyze_ping(result.stdout, host)
        return {
            'server': host,
            'sent': PING_COUNT,
            'received': stats['received'],
            'lost': stats['lost'],
            'packet_loss': (stats['lost']/PING_COUNT)*100,
            'average_ping': mean(stats['times']) if stats['times'] else 0,
            'status': 'Failed' if stats['received'] == 0 else 'Success',
            'response_time': f"{time.time()-start_time:.2f}s"
        }
    except subprocess.TimeoutExpired:
        debug_print(f"Timeout processing {host}")
        return error_result(host, "Timeout")
    except Exception as e:
        debug_print(f"Unexpected error pinging {host}: {str(e)}")
        return error_result(host, str(e))

def error_result(host, error):
    return {
        'server': host,
        'sent': PING_COUNT,
        'received': 0,
        'lost': PING_COUNT,
        'packet_loss': 100.0,
        'average_ping': 0,
        'status': f"Error: {error}",
        'response_time': 'N/A'
    }

def create_excel_report(results, file_prefix):
    wb = Workbook()
    ws = wb.active
    ws.title = file_prefix.upper()

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="4F81BD")
    good_fill = PatternFill("solid", fgColor="C6EFCE")
    bad_fill = PatternFill("solid", fgColor="FFC7CE")

    headers = ["Server", "Sent", "Received", "Lost", "Packet Loss %",
               "Avg Ping (ms)", "Status", "Response Time"]
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill

    for row, result in enumerate(results, 2):
        status = result['status']
        ws.cell(row=row, column=1, value=result['server'])
        ws.cell(row=row, column=2, value=result['sent'])
        ws.cell(row=row, column=3, value=result['received'])
        ws.cell(row=row, column=4, value=result['lost'])
        ws.cell(row=row, column=5, value=f"{result['packet_loss']:.1f}%" if 'Error' not in status else "N/A")
        ws.cell(row=row, column=6, value=f"{result['average_ping']:.2f}" if result['average_ping'] else "N/A")
        ws.cell(row=row, column=7, value=status)
        ws.cell(row=row, column=8, value=result['response_time'])

        row_fill = good_fill if status == "Success" else bad_fill
        for col in range(1, 9):
            ws.cell(row=row, column=col).fill = row_fill

    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['G'].width = 20
    ws.column_dimensions['H'].width = 15

    os.makedirs(RESULTS_DIR, exist_ok=True)
    output_file = os.path.join(RESULTS_DIR, f"{file_prefix}_results.xlsx")
    wb.save(output_file)
    debug_print(f"Report saved to {output_file}")
    return output_file

def validate_server_file(file_path):
    if not os.path.exists(file_path):
        debug_print(f"File not found: {file_path}")
        return False
    
    with open(file_path, 'r') as f:
        servers = [line.strip() for line in f if line.strip()]
    
    if not servers:
        debug_print(f"Empty server file: {file_path}")
        return False
    
    return True

def process_server_file(file_path, file_prefix):
    debug_print(f"Processing {file_prefix} from {file_path}")
    
    if not validate_server_file(file_path):
        return None
    
    with open(file_path, 'r') as f:
        servers = [line.strip() for line in f if line.strip()]
    
    results = []
    debug_print(f"Testing {len(servers)} servers from {file_prefix}")
    
    with ThreadPoolExecutor(max_workers=MAX_THREADS) as executor:
        futures = [executor.submit(ping_host, server) for server in servers]
        for idx, future in enumerate(as_completed(futures), 1):
            result = future.result()
            results.append(result)
            debug_print(f"Completed {idx}/{len(servers)}: {result['server']} ({result['status']})")
    
    results.sort(key=lambda x: (
        x['status'] != "Success",
        x['average_ping'] if x['average_ping'] else float('inf')
    ))
    
    return create_excel_report(results, file_prefix)

def main():
    print("\n=== Server Ping Tester ===\n")
    setup_dirs()
    
    if not os.listdir(SERVERS_DIR):
        print(f"No server files found in {SERVERS_DIR}")
        print("Add .txt files with one server per line")
        return
    
    results = {}
    for filename in os.listdir(SERVERS_DIR):
        if filename.endswith('.txt'):
            file_prefix = os.path.splitext(filename)[0]
            file_path = os.path.join(SERVERS_DIR, filename)
            report_file = process_server_file(file_path, file_prefix)
            if report_file:
                results[file_prefix] = report_file
    
    print("\n=== Testing Summary ===")
    for prefix, report in results.items():
        print(f"{prefix.upper()}: {report}")
    print("\nTesting complete!")

if __name__ == '__main__':
    main()
